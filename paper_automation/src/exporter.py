"""
엑셀 출력 모듈
수집된 논문 데이터를 Excel 파일로 출력
"""

import os
import logging
from typing import Dict, List, Optional
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel 파일 출력 클래스"""
    
    def __init__(self, output_path: str = "./results"):
        """초기화
        
        Args:
            output_path: 출력 경로
        """
        self.output_path = Path(output_path)
        self.output_path.mkdir(parents=True, exist_ok=True)
        
    def export_papers(self, papers: List[Dict], filename: Optional[str] = None) -> str:
        """논문 데이터를 Excel로 출력
        
        Args:
            papers: 논문 데이터 리스트
            filename: 출력 파일명
            
        Returns:
            생성된 파일 경로
        """
        if not papers:
            logger.warning("출력할 논문이 없습니다")
            return None
            
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"papers_{timestamp}.xlsx"
            
        filepath = self.output_path / filename
        
        try:
            # DataFrame 생성
            df = self._create_dataframe(papers)
            
            # Excel Writer 생성
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 메인 시트에 데이터 쓰기
                df.to_excel(writer, sheet_name='논문목록', index=False)
                
                # 스타일 적용
                workbook = writer.book
                worksheet = writer.sheets['논문목록']
                self._apply_styles(worksheet, len(df))
                
                # 통계 시트 추가
                stats_df = self._create_statistics_df(papers)
                stats_df.to_excel(writer, sheet_name='통계', index=False)
                stats_worksheet = writer.sheets['통계']
                self._apply_styles(stats_worksheet, len(stats_df))
                
            logger.info(f"Excel 파일 생성 완료: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Excel 출력 중 오류: {e}")
            raise
            
    def _create_dataframe(self, papers: List[Dict]) -> pd.DataFrame:
        """논문 데이터를 DataFrame으로 변환
        
        Args:
            papers: 논문 데이터 리스트
            
        Returns:
            DataFrame
        """
        # 컬럼 순서 정의
        columns = [
            '제목',
            '저자',
            '발행연도',
            '학술지',
            '키워드',
            '초록',
            'AI 요약',
            '원문 링크',
            '수집일시'
        ]
        
        # 데이터 정리
        data = []
        for paper in papers:
            row = {
                '제목': paper.get('title', ''),
                '저자': paper.get('authors', ''),
                '발행연도': paper.get('year', ''),
                '학술지': paper.get('journal', paper.get('publication', '')),
                '키워드': ', '.join(paper.get('keywords', [])) if isinstance(paper.get('keywords'), list) else paper.get('keywords', ''),
                '초록': paper.get('abstract', '')[:500] + '...' if len(paper.get('abstract', '')) > 500 else paper.get('abstract', ''),
                'AI 요약': paper.get('summary', ''),
                '원문 링크': paper.get('fulltext_link', paper.get('link', '')),
                '수집일시': paper.get('collected_at', datetime.now().isoformat())
            }
            data.append(row)
            
        df = pd.DataFrame(data, columns=columns)
        return df
        
    def _apply_styles(self, worksheet, row_count: int):
        """Excel 시트에 스타일 적용
        
        Args:
            worksheet: openpyxl worksheet
            row_count: 데이터 행 수
        """
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 행 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # 데이터 행 스타일 적용
        for row in worksheet.iter_rows(min_row=2, max_row=row_count+1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border
                
        # 컬럼 너비 조정
        column_widths = {
            'A': 40,  # 제목
            'B': 20,  # 저자
            'C': 10,  # 발행연도
            'D': 25,  # 학술지
            'E': 30,  # 키워드
            'F': 50,  # 초록
            'G': 50,  # AI 요약
            'H': 30,  # 원문 링크
            'I': 20,  # 수집일시
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
            
        # 행 높이 자동 조정
        for row in range(2, row_count + 2):
            worksheet.row_dimensions[row].height = 60
            
    def _create_statistics_df(self, papers: List[Dict]) -> pd.DataFrame:
        """통계 정보 DataFrame 생성
        
        Args:
            papers: 논문 데이터 리스트
            
        Returns:
            통계 DataFrame
        """
        stats = {
            '항목': [],
            '값': []
        }
        
        # 기본 통계
        stats['항목'].append('총 논문 수')
        stats['값'].append(len(papers))
        
        stats['항목'].append('초록 있는 논문')
        stats['값'].append(sum(1 for p in papers if p.get('abstract')))
        
        stats['항목'].append('원문 링크 있는 논문')
        stats['값'].append(sum(1 for p in papers if p.get('fulltext_link') or p.get('link')))
        
        stats['항목'].append('AI 요약 생성된 논문')
        stats['값'].append(sum(1 for p in papers if p.get('summary')))
        
        # 연도별 분포
        years = {}
        for paper in papers:
            year = paper.get('year', 'Unknown')
            years[year] = years.get(year, 0) + 1
            
        for year, count in sorted(years.items()):
            stats['항목'].append(f'{year}년 논문')
            stats['값'].append(count)
            
        return pd.DataFrame(stats)
        
    def export_summary_report(self, papers: List[Dict], executive_summary: str, filename: Optional[str] = None) -> str:
        """종합 보고서 생성
        
        Args:
            papers: 논문 데이터 리스트
            executive_summary: 종합 요약
            filename: 출력 파일명
            
        Returns:
            생성된 파일 경로
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{timestamp}.xlsx"
            
        filepath = self.output_path / filename
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 요약 시트
                summary_df = pd.DataFrame({
                    '종합 분석': [executive_summary]
                })
                summary_df.to_excel(writer, sheet_name='종합요약', index=False)
                
                # 논문 목록 시트
                papers_df = self._create_dataframe(papers)
                papers_df.to_excel(writer, sheet_name='논문목록', index=False)
                
                # 통계 시트
                stats_df = self._create_statistics_df(papers)
                stats_df.to_excel(writer, sheet_name='통계', index=False)
                
                # 스타일 적용
                workbook = writer.book
                
                # 종합요약 시트 스타일
                summary_sheet = writer.sheets['종합요약']
                summary_sheet.column_dimensions['A'].width = 100
                for cell in summary_sheet['A']:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                # 논문목록 시트 스타일
                papers_sheet = writer.sheets['논문목록']
                self._apply_styles(papers_sheet, len(papers_df))
                
                # 통계 시트 스타일
                stats_sheet = writer.sheets['통계']
                self._apply_styles(stats_sheet, len(stats_df))
                
            logger.info(f"종합 보고서 생성 완료: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"보고서 생성 중 오류: {e}")
            raise
            
    def merge_excel_files(self, filepaths: List[str], output_filename: str) -> str:
        """여러 Excel 파일 병합
        
        Args:
            filepaths: 병합할 파일 경로 리스트
            output_filename: 출력 파일명
            
        Returns:
            병합된 파일 경로
        """
        all_papers = []
        
        for filepath in filepaths:
            try:
                df = pd.read_excel(filepath, sheet_name='논문목록')
                papers = df.to_dict('records')
                all_papers.extend(papers)
            except Exception as e:
                logger.warning(f"파일 읽기 실패 {filepath}: {e}")
                continue
                
        # 중복 제거 (제목 기준)
        unique_papers = []
        seen_titles = set()
        
        for paper in all_papers:
            title = paper.get('제목', '')
            if title and title not in seen_titles:
                unique_papers.append(paper)
                seen_titles.add(title)
                
        # DataFrame으로 변환하고 저장
        merged_df = pd.DataFrame(unique_papers)
        output_path = self.output_path / output_filename
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='병합된_논문목록', index=False)
            worksheet = writer.sheets['병합된_논문목록']
            self._apply_styles(worksheet, len(merged_df))
            
        logger.info(f"파일 병합 완료: {output_path} ({len(unique_papers)}개 논문)")
        return str(output_path)


# 테스트용 함수
def test_exporter():
    """출력기 테스트"""
    exporter = ExcelExporter()
    
    # 샘플 데이터
    sample_papers = [
        {
            "title": "AI와 머신러닝 연구",
            "authors": "홍길동, 김철수",
            "abstract": "이 논문은 AI와 머신러닝에 대한 최신 연구 동향을 다룹니다.",
            "year": "2024",
            "journal": "한국AI학회지",
            "keywords": ["AI", "머신러닝", "딥러닝"],
            "summary": "이 논문은 AI 기술의 발전과 응용에 대해 논의합니다.",
            "fulltext_link": "https://example.com/paper1.pdf",
            "collected_at": datetime.now().isoformat()
        },
        {
            "title": "자연어 처리 기술 동향",
            "authors": "이영희",
            "abstract": "최신 자연어 처리 기술과 그 응용 사례를 소개합니다.",
            "year": "2024",
            "journal": "컴퓨터공학회지",
            "keywords": ["NLP", "BERT", "GPT"],
            "summary": "자연어 처리 분야의 최신 트렌드를 분석했습니다.",
            "fulltext_link": "https://example.com/paper2.pdf",
            "collected_at": datetime.now().isoformat()
        }
    ]
    
    # Excel 출력
    filepath = exporter.export_papers(sample_papers)
    print(f"Excel 파일 생성: {filepath}")
    
    # 종합 보고서 생성
    executive_summary = "수집된 논문들은 주로 AI와 자연어 처리 분야를 다루고 있으며..."
    report_path = exporter.export_summary_report(sample_papers, executive_summary)
    print(f"종합 보고서 생성: {report_path}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    test_exporter()